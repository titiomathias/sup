# Código de Execução Back-end
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font

# Criando classe de criação back-end
class CriarPlanilha:

    # Recebendo dados das interfaces gráficas.
    def __init__(self, nome_planilha, data, client, cnpj, olt, ont, sn, v_firmware, v_hardware, duvida, resolucao, atendimento, local):
        nome_planilha = nome_planilha + ".xlsx"

        # Criando documento.
        planilha = openpyxl.Workbook()
        planilha.create_sheet("Report")
        pagina = planilha["Report"]

        # Personalizando documento.
        pagina.append(["Data", "Cliente", "CNPJ", "Modelo da OlT", "Modelo do Equipamento", "Serial Number", "Versão de Firmware", "Versão de Hardware", "Dúvida/Problema", "Solução", "Atendimento"])
        pagina["A1"].font = Font(bold=True)
        pagina["B1"].font = Font(bold=True)
        pagina["C1"].font = Font(bold=True)
        pagina["D1"].font = Font(bold=True)
        pagina["E1"].font = Font(bold=True)
        pagina["F1"].font = Font(bold=True)
        pagina["G1"].font = Font(bold=True)
        pagina["H1"].font = Font(bold=True)
        pagina["I1"].font = Font(bold=True)
        pagina["J1"].font = Font(bold=True)
        pagina["K1"].font = Font(bold=True)

        pagina.column_dimensions['A'].width = 15
        pagina.column_dimensions['B'].width = 15
        pagina.column_dimensions['C'].width = 18
        pagina.column_dimensions['D'].width = 24
        pagina.column_dimensions['E'].width = 24
        pagina.column_dimensions['E'].width = 24
        pagina.column_dimensions['F'].width = 24
        pagina.column_dimensions['G'].width = 24
        pagina.column_dimensions['H'].width = 24
        pagina.column_dimensions['I'].width = 36
        pagina.column_dimensions['J'].width = 36
        pagina.column_dimensions['K'].width = 36

        # Inserindo data
        for row in pagina["A2:A100"]:
            for cell in row:
                retorno = True
                if cell.value == None:
                    pagina.cell(row=cell.row, column=cell.column).value = data
                    retorno = False
                    break
                else:
                    pass
            if retorno == False:
                break
            else:
                continue


        # Inserindo o nome do cliente
        for row in pagina["B2:B100"]:
            for cell in row:
                retorno = True
                if cell.value == None:
                    pagina.cell(row=cell.row, column=cell.column).value = client
                    retorno = False
                    break
                else:
                    pass
            if retorno == False:
                break
            else:
                continue

        # Inserindo cnpj do cliente
        for row in pagina["C2:C100"]:
            for cell in row:
                retorno = True
                if cell.value == None:
                    pagina.cell(row=cell.row, column=cell.column).value = cnpj
                    retorno = False
                    break
                else:
                    pass
            if retorno == False:
                break
            else:
                continue

        # Inserindo modelo da olt do cliente
        for row in pagina["D2:D100"]:
            for cell in row:
                retorno = True
                if cell.value == None:
                    pagina.cell(row=cell.row, column=cell.column).value = olt
                    retorno = False
                    break
                else:
                    pass
            if retorno == False:
                break
            else:
                continue

        # Inserindo equipamento do cliente
        for row in pagina["E2:E100"]:
            for cell in row:
                retorno = True
                if cell.value == None:
                    pagina.cell(row=cell.row, column=cell.column).value = ont
                    retorno = False
                    break
                else:
                    pass
            if retorno == False:
                break
            else:
                continue

        # Serial Number do equipamento
        for row in pagina["F2:F100"]:
            for cell in row:
                retorno = True
                if cell.value == None:
                    pagina.cell(row=cell.row, column=cell.column).value = sn
                    retorno = False
                    break
                else:
                    pass
            if retorno == False:
                break
            else:
                continue

        # Versão do firmware do equipamento
        for row in pagina["G2:G100"]:
            for cell in row:
                retorno = True
                if cell.value == None:
                    pagina.cell(row=cell.row, column=cell.column).value = v_firmware
                    retorno = False
                    break
                else:
                    pass
            if retorno == False:
                break
            else:
                continue

        # Versão do hardware do equipamento
        for row in pagina["H2:H100"]:
            for cell in row:
                retorno = True
                if cell.value == None:
                    pagina.cell(row=cell.row, column=cell.column).value = v_hardware
                    retorno = False
                    break
                else:
                    pass
            if retorno == False:
                break
            else:
                continue


        # Duvida do cliente
        for row in pagina["I2:I100"]:
            for cell in row:
                retorno = True
                if cell.value == None:
                    pagina.cell(row=cell.row, column=cell.column).value = duvida
                    retorno = False
                    break
                else:
                    pass
            if retorno == False:
                break
            else:
                continue

        # Conclusão de atendimento: Bem-sucedido ou mal-sucedido
        for row in pagina["K2:K100"]:
            for cell in row:
                retorno = True
                if cell.value == None:
                    if atendimento == True:
                        pagina.cell(row=cell.row, column=cell.column).value = "Atendimento bem sucedido."
                    else:
                        pagina.cell(row=cell.row, column=cell.column).value = "Atendimento mal sucedido."
                    retorno = False
                    break
                else:
                    pass
            if retorno == False:
                break
            else:
                continue

        if atendimento == True:
            # resolucao
            for row in pagina["J2:J100"]:
                for cell in row:
                    retorno = True
                    if cell.value == None:
                        pagina.cell(row=cell.row, column=cell.column).value = resolucao
                        retorno = False
                        break
                    else:
                        pass
                if retorno == False:
                    break
                else:
                    continue
        else:
            resolucao = "Problema não solucionado."
            for row in pagina["J2:J100"]:
                for cell in row:
                    retorno = True
                    if cell.value == None:
                        pagina.cell(row=cell.row, column=cell.column).value = resolucao
                        retorno = False
                        break
                    else:
                        pass
                if retorno == False:
                    break
                else:
                    continue

        planilha.save(filename=f"{local}{nome_planilha}")

# Criando classe de edição back-end
class EditarPlanilha:

    # Recebendo dados das interfaces gráficas.
    def __init__(self, nome_planilha, data, client, cnpj, olt, ont, sn, v_firmware, v_hardware, duvida, resolucao, atendimento):

        # Abrindo documento para edição.
        planilha = load_workbook(filename=f"{nome_planilha}")
        pagina = planilha["Report"]

        # Inserindo data
        for row in pagina["A2:A100"]:
            for cell in row:
                retorno = True
                if cell.value == None:
                    pagina.cell(row=cell.row, column=cell.column).value = data
                    retorno = False
                    break
                else:
                    pass
            if retorno == False:
                break
            else:
                continue

        # Inserindo nome do cliente
        for row in pagina["B2:B100"]:
            for cell in row:
                retorno = True
                if cell.value == None:
                    pagina.cell(row=cell.row, column=cell.column).value = client
                    retorno = False
                    break
                else:
                    pass
            if retorno == False:
                break
            else:
                continue

        # Inserindo cnpj do cliente
        for row in pagina["C2:C100"]:
            for cell in row:
                retorno = True
                if cell.value == None:
                    pagina.cell(row=cell.row, column=cell.column).value = cnpj
                    retorno = False
                    break
                else:
                    pass
            if retorno == False:
                break
            else:
                continue

        # Inserindo modelo olt do cliente
        for row in pagina["D2:D100"]:
            for cell in row:
                retorno = True
                if cell.value == None:
                    pagina.cell(row=cell.row, column=cell.column).value = olt
                    retorno = False
                    break
                else:
                    pass
            if retorno == False:
                break
            else:
                continue

        # Inserindo equipamento usado pelo cliente
        for row in pagina["E2:E100"]:
            for cell in row:
                retorno = True
                if cell.value == None:
                    pagina.cell(row=cell.row, column=cell.column).value = ont
                    retorno = False
                    break
                else:
                    pass
            if retorno == False:
                break
            else:
                continue

        # Inserindo Serial Number do equipamento
        for row in pagina["F2:F100"]:
            for cell in row:
                retorno = True
                if cell.value == None:
                    pagina.cell(row=cell.row, column=cell.column).value = sn
                    retorno = False
                    break
                else:
                    pass
            if retorno == False:
                break
            else:
                continue

        # Inserindo versão do firmware do equipamento
        for row in pagina["G2:G100"]:
            for cell in row:
                retorno = True
                if cell.value == None:
                    pagina.cell(row=cell.row, column=cell.column).value = v_firmware
                    retorno = False
                    break
                else:
                    pass
            if retorno == False:
                break
            else:
                continue

        # Inserindo versão do hardware do equipamento
        for row in pagina["H2:H100"]:
            for cell in row:
                retorno = True
                if cell.value == None:
                    pagina.cell(row=cell.row, column=cell.column).value = v_hardware
                    retorno = False
                    break
                else:
                    pass
            if retorno == False:
                break
            else:
                continue

        # Inserindo dúvida do cliente
        for row in pagina["I2:I100"]:
            for cell in row:
                retorno = True
                if cell.value == None:
                    pagina.cell(row=cell.row, column=cell.column).value = duvida
                    retorno = False
                    break
                else:
                    pass
            if retorno == False:
                break
            else:
                continue

        # Inserindo conclusão de atendimento: Bem sucedido ou mal sucedido
        for row in pagina["K2:K100"]:
            for cell in row:
                retorno = True
                if cell.value == None:
                    if atendimento == True:
                        pagina.cell(row=cell.row, column=cell.column).value = "Atendimento bem sucedido."
                    else:
                        pagina.cell(row=cell.row, column=cell.column).value = "Atendimento mal sucedido."
                    retorno = False
                    break
                else:
                    pass
            if retorno == False:
                break
            else:
                continue

        if atendimento == True:
            # resolucao
            for row in pagina["J2:J100"]:
                for cell in row:
                    retorno = True
                    if cell.value == None:
                        pagina.cell(row=cell.row, column=cell.column).value = resolucao
                        retorno = False
                        break
                    else:
                        pass
                if retorno == False:
                    break
                else:
                    continue
        else:
            resolucao = "Problema não solucionado."
            for row in pagina["J2:J100"]:
                for cell in row:
                    retorno = True
                    if cell.value == None:
                        pagina.cell(row=cell.row, column=cell.column).value = resolucao
                        retorno = False
                        break
                    else:
                        pass
                if retorno == False:
                    break
                else:
                    continue

        planilha.save(filename=f"{nome_planilha}")